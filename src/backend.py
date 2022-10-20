from numpy import diff
import pandas as pd
import os.path
import warnings
import holidays
import calendar
from datetime import datetime, date, timedelta
from openpyxl import load_workbook

COUNTRY = 'DE'
STATE = 'NW'
ANNUAL_LEAVE_DAYS = 30
WEEKLY_WORKING_HOURS = 40
WEEKLY_WORKING_DAYS = 5
DAILY_WORKING_HOURS = int(WEEKLY_WORKING_HOURS / WEEKLY_WORKING_DAYS)

def sum_time(t1: str, t2: str) -> str:
    """
    Sum up two time, return the sum in formatted string 'HH:MM'
    Input parameters in string format 'HH:MM'
    """
    t1 += ':00'
    t2 += ':00'

    t1 = pd.to_timedelta(t1)
    t2 = pd.to_timedelta(t2)

    t_sum = t1 + t2         # timedelta
    t_sum = str(t_sum).split(' ')[2][:-3]   # formatted 'HH:MM'

    return t_sum 

def compare_time(t1: str, t2: str) -> bool:
    """
    Compare between two time, return True if t1 ahead of t2, otherwise return False
    Input parameters in string format 'HH:MM'
    """
    if (pd.to_datetime(t1) - pd.to_datetime(t2)).total_seconds() > 0:
        return True
    else:
        return False

def calc_duration(begin_t: str, end_t: str) -> str:
    """
    Calculates the time duration, return the result as string format 'HH:MM'
    Input parameters in string format 'HH:MM'
    """
    end_t = pd.to_datetime(end_t)
    begin_t = pd.to_datetime(begin_t)

    if (end_t - begin_t).total_seconds() > 0:
        duration = (end_t - begin_t + timedelta(hours=24)
                    ) % timedelta(hours=24)
        duration = str(duration).split(' ')[2]
    else:
        warnings.warn('Warning: Invalid input, end time before begin time')
        return

    return duration[:-3]

def calc_worksum(df:pd.DataFrame, day=1) -> str:
    """
    Calculates the total working hours till the given day in a month, return in formmatted string 'HH:MM'
    """
    df_update = df.loc[:day-1, 'Work Sum'] + ':00'
    df_update = pd.to_timedelta(df_update)
    actual_working_sum = df_update.sum()        # timedelta object

    hours = actual_working_sum.seconds // 3600
    minutes = actual_working_sum.seconds // 60 - (hours * 60)
    hours += actual_working_sum.days * 24 

    return '{:02d}:{:02d}'.format(hours, minutes)

def calc_overtime(df:pd.DataFrame, day=1) -> str:
    """
    Calculates the overtime till the given day in a month, return in formmatted string 'HH:MM'
    """
    actual_working_sum = df.loc[df['Date'] == 'Summary', 'Work Sum'].values[0]
    actual_working_in_min = 60 * int(actual_working_sum[:2]) + int(actual_working_sum[-2:])
    
    num_working_day = df.loc[:day-1, 'Comment'].value_counts()['working day']
    target_working_sum_in_min = num_working_day * DAILY_WORKING_HOURS * 60

    diff_in_min = actual_working_in_min - target_working_sum_in_min
    diff_hour = abs(diff_in_min) // 60
    diff_minute = abs(diff_in_min) - (diff_hour * 60)
    if diff_in_min < 0:
        return ('-{:2d}:{:2d}'.format(diff_hour, diff_minute))
    else:
        return ('{:2d}:{:2d}'.format(diff_hour, diff_minute))

def create_month_report(year=2022, month=1) -> pd.DataFrame:
    """
    Create a monthly report dataframe
    """
    begin_date = '{}/1/{}'.format(month, year)      # MM/DD/YY
    days = calendar.monthrange(year, month)[1]      # get total days of month
    month_range = pd.date_range(start=begin_date, periods=days)

    public_holidays = holidays.country_holidays(COUNTRY, STATE, year)

    df = pd.DataFrame({'Date': month_range.strftime('%Y-%m-%d'),
                       'Day': month_range.strftime('%a'),
                       'Clock In': pd.Series([' : '] * days),
                       'Clock Out': pd.Series([' : '] * days),
                       'Pause Start': pd.Series([' : '] * days),
                       'Pause Stop': pd.Series([' : '] * days),
                       'Pause': pd.Series(['00:00'] * days),
                       'Work Sum': pd.Series(['00:00'] * days)})

    # Fill up col['Comment']
    for i in range(days):
        day_i = date(year, month, i+1)
        if day_i in public_holidays:
            df.loc[i, 'Comment'] = 'holiday ({})'.format(
                public_holidays[day_i])
        else:
            if day_i.weekday() >= WEEKLY_WORKING_DAYS:
                df.loc[i, 'Comment'] = 'not working day'
            else:
                df.loc[i, 'Comment'] = 'working day'

    # Append last row for summary
    df.loc[days, 'Date'] = 'Summary'
    df.loc[days, 'Pause'] = '00:00'
    df.loc[days, 'Work Sum'] = '00:00'
    df.fillna(' ', inplace=True)

    return df


def save_month_report(df: pd.DataFrame, year=2022, month=1):
    """
    save the month report dataframe as an Excel file,
    file name is distiguished by year, and month report are saved in different sheets.
    """
    f_name = './docs/' + 'attendance_' + str(year) + '.xlsx'

    # if file not exists, save it to the working dir
    if not os.path.exists(f_name):
        df.to_excel(f_name, index=False, sheet_name=str(month))
    # if file already exists, append the new month report to a separate sheet
    else:
        with pd.ExcelWriter(f_name, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=str(month))

    return


def load_month_report(year=2022, month=1) -> pd.DataFrame:
    """
    Loads the month report as a dataframe from excel sheet
    """
    f_name = './docs/' + 'attendance_' + str(year) + '.xlsx'

    # if file not exists, create a new and save
    if not os.path.exists(f_name):
        df = create_month_report(year, month)
        save_month_report(df, year, month)
        return df

    # if file exists, but sheet not exists, save as a new sheet
    wb = load_workbook(f_name)
    if not str(month) in wb.sheetnames:
        df = create_month_report(year, month)
        save_month_report(df, year, month)
        return df

    # return month report dataframe if already exists in the file
    df = pd.read_excel(f_name, sheet_name=str(month), index_col=False)

    return df


def load_year_report(year=2022) -> pd.DataFrame:
    """
    Loads the yearly report as a dataframe from excel sheet
    """
    f_name = './docs/' + 'attendance_' + str(year) + '.xlsx'

    df = pd.DataFrame()

    xls = pd.ExcelFile(f_name)
    for sheet in xls.sheet_names:
        df_sheet = pd.read_excel(f_name, sheet_name=sheet)
        df = pd.concat([df, df_sheet], ignore_index=True)

    return df
